import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

df_global = None

def load_csv_to_xlsx_file(csv_file_path, excel_file_path, separator=';', encoding='utf-8'):
    global df_global
    try:
        print("Carregando arquivo CSV...")
        df_global = pd.read_csv(csv_file_path, sep=separator, encoding=encoding)
        print("Arquivo CSV carregado com sucesso.")
        
        df_global.to_excel(excel_file_path, index=False, sheet_name='Principal')
        print("DataFrame salvo no arquivo Excel.")
        
        adjust_column_width(excel_file_path, 'Principal')
    except UnicodeDecodeError as e:
        print(f"Erro de decodificação de caracteres: {e}")
    except Exception as e:
        print(f"Erro ao carregar o arquivo CSV: {e}")

def filter_value_by_column_name(column: str, values: list):
    global df_global
    print(f"Filtrando valores pela coluna: {column}")
    return df_global[df_global[column].isin(values)]

def save_new_sheet(df_data, new_tab_name):
    global excel_file_path
    print(f"Criando Planilha {new_tab_name}")
    try:
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:

            book = load_workbook(excel_file_path)
            if new_tab_name in book.sheetnames:
                existing_df = pd.read_excel(excel_file_path, sheet_name=new_tab_name)
                df_data = pd.concat([existing_df, df_data], ignore_index=True)

                del book[new_tab_name]
                writer.book = book

            df_data.to_excel(writer, sheet_name=new_tab_name, index=False)
            print(f"Dados salvos na nova planilha: {new_tab_name}")
            
            adjust_column_width(excel_file_path, new_tab_name)
    except Exception as e:
        print(f"Erro ao salvar o DataFrame no arquivo Excel: {e}")

def save_and_filter_sheet_by_obj(column, obj_sheet: object):
    for key, values in obj_sheet.items():
        sheet_data = filter_value_by_column_name(column, values)

        sheet_name = f"{key}"
        save_new_sheet(sheet_data, sheet_name)

def adjust_column_width(excel_file_path, sheet_name):
    try:
        book = load_workbook(excel_file_path)
        sheet = book[sheet_name]

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1)
            sheet.column_dimensions[column].width = adjusted_width

        book.save(excel_file_path)
        print(f"Largura das colunas ajustada para a planilha: {sheet_name}")
    except Exception as e:
        print(f"Erro ao ajustar a largura das colunas: {e}")

def main():
    try:
        with open('config.json', 'r', encoding='utf-8') as config_file:
            config = json.load(config_file)
    except Exception as e:
        print(f"Erro ao carregar o arquivo de configuração: {e}")
        return

    #csv_file_path = input("Por favor, insira o caminho do arquivo CSV: ")
    csv_file_path = config.get("csvFilePath", "Todos os Processos.csv")
    global excel_file_path
    excel_file_path = config.get("excelFilePath", "FilteredSpreadsheet.xlsx")
    separator = config.get("separator", ";")
    encoding = config.get("encoding", "utf-8") 
    movements = config.get("valoresParaFiltagrem", {})
    filter_column_name = config.get("nomeColunaFiltro")

    print("Carregando Planilhas...")
    load_csv_to_xlsx_file(csv_file_path, excel_file_path, separator, encoding)
    print("Fazendo Filtragem...")
    save_and_filter_sheet_by_obj(filter_column_name, movements)

if __name__ == "__main__":
    main()
    print("Execução do script concluída.")